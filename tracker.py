"""
Application tracker: logs every job you apply to into an Excel file, and
helps you keep statuses up to date over time.

This deliberately doesn't try to auto-detect application status by
logging into LinkedIn/Naukri -- there's no safe way to automate that
(see README). Instead, updating status is a 10-second command you run
yourself after checking your applications.

Usage:
    # Log a new application (do this right after you apply)
    python tracker.py add --title "Backend Engineer" --company "Fairground" \
        --platform LinkedIn --url "https://..." --score 0.68

    # See everything you've applied to
    python tracker.py list

    # See applications that haven't been updated in a while and need a status check
    python tracker.py stale --days 14

    # Update the status of an application (interactive)
    python tracker.py update
"""

import argparse
import os
import sys
from datetime import datetime, timedelta

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

XLSX_PATH = "applications.xlsx"
SHEET_NAME = "Applications"

HEADERS = [
    "Date Applied", "Job Title", "Company", "Platform", "Match Score",
    "Job URL", "Status", "Last Updated", "Notes",
]
COL_WIDTHS = [14, 32, 24, 12, 12, 42, 14, 14, 30]

VALID_STATUSES = [
    "Applied", "Viewed", "Interview", "Rejected", "Offer", "No Response", "Withdrawn",
]


def _ensure_workbook():
    if os.path.exists(XLSX_PATH):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    wb.save(XLSX_PATH)


def _load_sheet():
    _ensure_workbook()
    wb = load_workbook(XLSX_PATH)
    return wb, wb[SHEET_NAME]


def _all_rows(ws):
    """Yield (row_index, dict-of-values) for every data row."""
    for row_idx in range(2, ws.max_row + 1):
        values = [ws.cell(row=row_idx, column=c).value for c in range(1, len(HEADERS) + 1)]
        if all(v is None for v in values):
            continue
        yield row_idx, dict(zip(HEADERS, values))


def add_application(title, company, platform, url="", score=None, notes=""):
    wb, ws = _load_sheet()

    # Avoid duplicate logging of the same job
    for _, row in _all_rows(ws):
        if (row["Job Title"] or "").strip().lower() == title.strip().lower() and \
           (row["Company"] or "").strip().lower() == company.strip().lower() and \
           (row["Platform"] or "").strip().lower() == platform.strip().lower():
            print(f"Already logged: '{title}' at '{company}' via {platform} -- skipping duplicate.")
            return

    today = datetime.now().strftime("%Y-%m-%d")
    new_row = [
        today, title, company, platform,
        round(score, 4) if score is not None else "",
        url, "Applied", today, notes,
    ]
    font = Font(name="Arial")
    row_idx = ws.max_row + 1
    for col_idx, value in enumerate(new_row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = font

    wb.save(XLSX_PATH)
    print(f"Logged: '{title}' at '{company}' via {platform}")


def list_applications():
    _, ws = _load_sheet()
    rows = list(_all_rows(ws))
    if not rows:
        print("No applications logged yet. Use: python tracker.py add --title ... --company ... --platform ...")
        return

    print(f"{'#':<4}{'Date':<12}{'Title':<32}{'Company':<22}{'Platform':<12}{'Status':<12}")
    print("-" * 94)
    for row_idx, row in rows:
        print(f"{row_idx-1:<4}{str(row['Date Applied']):<12}{str(row['Job Title'])[:30]:<32}"
              f"{str(row['Company'])[:20]:<22}{str(row['Platform']):<12}{str(row['Status']):<12}")


def show_stale(days=14):
    _, ws = _load_sheet()
    cutoff = datetime.now() - timedelta(days=days)
    stale = []
    for row_idx, row in _all_rows(ws):
        if row["Status"] not in ("Applied", "Viewed"):
            continue
        try:
            last_updated = datetime.strptime(str(row["Last Updated"]), "%Y-%m-%d")
        except (ValueError, TypeError):
            continue
        if last_updated <= cutoff:
            stale.append((row_idx, row))

    if not stale:
        print(f"Nothing stale -- no open applications older than {days} days.")
        return

    print(f"{len(stale)} application(s) haven't been updated in {days}+ days. Worth checking status:\n")
    for row_idx, row in stale:
        print(f"  #{row_idx-1}  {row['Job Title']} @ {row['Company']} ({row['Platform']}) "
              f"-- applied {row['Date Applied']}, still '{row['Status']}'")
    print("\nRun: python tracker.py update")


def update_status_interactive():
    wb, ws = _load_sheet()
    rows = [(idx, row) for idx, row in _all_rows(ws) if row["Status"] not in ("Rejected", "Withdrawn")]
    if not rows:
        print("No open applications to update.")
        return

    print("Open applications:\n")
    for row_idx, row in rows:
        print(f"  #{row_idx-1}  {row['Job Title']} @ {row['Company']} ({row['Platform']}) -- currently '{row['Status']}'")

    choice = input("\nEnter # to update (or blank to cancel): ").strip()
    if not choice:
        return
    try:
        target_row = int(choice) + 1
    except ValueError:
        print("Not a valid number.")
        return

    print(f"\nValid statuses: {', '.join(VALID_STATUSES)}")
    new_status = input("New status: ").strip()
    if new_status not in VALID_STATUSES:
        print(f"'{new_status}' isn't in the known list -- saving it anyway.")

    status_col = HEADERS.index("Status") + 1
    updated_col = HEADERS.index("Last Updated") + 1
    ws.cell(row=target_row, column=status_col, value=new_status)
    ws.cell(row=target_row, column=updated_col, value=datetime.now().strftime("%Y-%m-%d"))
    wb.save(XLSX_PATH)
    print("Updated.")


def main():
    parser = argparse.ArgumentParser(description="Track job applications in an Excel file.")
    sub = parser.add_subparsers(dest="command", required=True)

    p_add = sub.add_parser("add", help="Log a new application")
    p_add.add_argument("--title", required=True)
    p_add.add_argument("--company", required=True)
    p_add.add_argument("--platform", required=True, help="e.g. LinkedIn, Naukri, Indeed, Company Site")
    p_add.add_argument("--url", default="")
    p_add.add_argument("--score", type=float, default=None, help="Match score from main.py, if known")
    p_add.add_argument("--notes", default="")

    sub.add_parser("list", help="List all logged applications")

    p_stale = sub.add_parser("stale", help="Show applications that need a status check")
    p_stale.add_argument("--days", type=int, default=14)

    sub.add_parser("update", help="Interactively update an application's status")

    args = parser.parse_args()

    if args.command == "add":
        add_application(args.title, args.company, args.platform, args.url, args.score, args.notes)
    elif args.command == "list":
        list_applications()
    elif args.command == "stale":
        show_stale(args.days)
    elif args.command == "update":
        update_status_interactive()


if __name__ == "__main__":
    main()
